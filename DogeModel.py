import requests
import pandas as pd
import numpy as np
import yfinance as yf
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_squared_error, r2_score
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import threading
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# Initialize sentiment analyzer
analyzer = SentimentIntensityAnalyzer()

def get_price_data(symbol, period='1y', interval='1d'):
    """
    Fetch historical price data for the given symbol using yfinance.
    """
    try:
        df = yf.download(symbol, period=period, interval=interval)
        df.reset_index(inplace=True)
        df.rename(columns={ 
            'Date': 'timestamp', 
            'Open': 'open', 
            'High': 'high', 
            'Low': 'low', 
            'Close': 'close', 
            'Volume': 'volume'
        }, inplace=True)
        return df
    except Exception as e:
        print(f"Error fetching data for {symbol}: {e}")
        return pd.DataFrame()

def get_sentiment_score(coin):
    """
    Analyze the sentiment of the given coin using Vader Sentiment Analyzer.
    """
    sentiment = analyzer.polarity_scores(coin)
    return sentiment['compound']

def get_on_chain_metrics(symbol):
    """
    Retrieve on-chain metrics for the given cryptocurrency symbol.
    """
    return {
        'transaction_count': np.random.randint(100, 1000),
        'active_addresses': np.random.randint(100, 1000)
    }

def calculate_rsi(close, window=14):
    """
    Calculate the Relative Strength Index (RSI) for a given series of closing prices.
    """
    delta = close.diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=window).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=window).mean()
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))
    return rsi

def calculate_macd(close):
    """
    Calculate the Moving Average Convergence Divergence (MACD) and its signal line.
    """
    macd = close.ewm(span=12, adjust=False).mean() - close.ewm(span=26, adjust=False).mean()
    macdsignal = macd.ewm(span=9, adjust=False).mean()
    return macd, macdsignal

def calculate_stochastic_rsi(rsi, window=14):
    """
    Calculate the Stochastic RSI for a given RSI series.
    """
    stoch_rsi = (rsi - rsi.rolling(window=window).min()) / (rsi.rolling(window=window).max() - rsi.rolling(window=window).min())
    return stoch_rsi

def calculate_technical_indicators(df):
    """
    Calculate technical indicators (RSI, MACD, Bollinger Bands, Stochastic RSI) for the given dataframe.
    """
    # Relative Strength Index (RSI)
    df['rsi'] = calculate_rsi(df['close'])
    
    # Moving Average Convergence Divergence (MACD)
    df['macd'], df['macdsignal'] = calculate_macd(df['close'])
    
    # Bollinger Bands
    df['bollinger_mavg'] = df['close'].rolling(window=20).mean()
    df['bollinger_std'] = df['close'].rolling(window=20).std()
    df['bollinger_upper'] = df['bollinger_mavg'] + 2 * df['bollinger_std']
    df['bollinger_lower'] = df['bollinger_mavg'] - 2 * df['bollinger_std']

    # Stochastic RSI
    df['stoch_rsi'] = calculate_stochastic_rsi(df['rsi'])
    
    return df

def calculate_scores(df, sentiment_score, on_chain_metrics):
    """
    Calculate final scores and recommendations based on technical indicators, sentiment scores, and on-chain metrics.
    """
    df.dropna(inplace=True)
    if not df.empty:
        scaler = MinMaxScaler()
        df[['rsi', 'macd']] = scaler.fit_transform(df[['rsi', 'macd']])
        
        on_chain_score = scaler.fit_transform(np.array(list(on_chain_metrics.values())).reshape(1, -1)).flatten()
        df['on_chain_score'] = on_chain_score[0]
        df['sentiment_score'] = sentiment_score

        df['final_score'] = df['rsi'] * 0.3 + df['macd'] * 0.3 + df['on_chain_score'] * 0.2 + df['sentiment_score'] * 0.2

        df['recommendation'] = df['final_score'].apply(lambda x: 'Buy' if x > 0.7 else 'Sell' if x < 0.3 else 'Hold')
    else:
        print("Empty DataFrame, cannot calculate scores.")

    return df

def preprocess_data(df):
    """
    Preprocess the data for predictive modeling.
    """
    df = df.dropna()
    df['timestamp'] = pd.to_datetime(df['timestamp']).astype(int) / 10**9  # Convert timestamp to seconds
    return df

def create_features(df):
    """
    Create features for the predictive model.
    """
    features = df[['timestamp', 'close', 'rsi', 'macd', 'macdsignal', 'bollinger_upper', 'bollinger_lower', 'sentiment_score', 'on_chain_score']]
    return features

def train_predictive_model(df):
    """
    Train a predictive model and make predictions.
    """
    df = preprocess_data(df)
    features = create_features(df)
    target = df['close']

    X_train, X_test, y_train, y_test = train_test_split(features, target, test_size=0.2, random_state=42)
    model = RandomForestRegressor(n_estimators=100, random_state=42)
    model.fit(X_train, y_train)

    y_pred = model.predict(X_test)
    mse = mean_squared_error(y_test, y_pred)
    r2 = r2_score(y_test, y_pred)

    print(f"Model Performance: MSE = {mse:.4f}, R^2 = {r2:.4f}")

    return model

def make_predictions(model, df, future_periods):
    """
    Make predictions using the trained model.
    """
    df = preprocess_data(df)
    features = create_features(df)
    predictions = model.predict(features)
    df['predicted_close'] = predictions
    
    # Generate future timestamps
    last_timestamp = df['timestamp'].iloc[-1]
    future_timestamps = [last_timestamp + (i * 60 * 60 * 24) for i in range(1, future_periods + 1)]
    
    # Create a new DataFrame for future predictions
    future_features = pd.DataFrame({
        'timestamp': future_timestamps,
        'close': df['close'].iloc[-1],  # Use the last close price as a starting point
        'rsi': df['rsi'].iloc[-1],  # Use the last RSI value as a starting point
        'macd': df['macd'].iloc[-1],  # Use the last MACD value as a starting point
        'macdsignal': df['macdsignal'].iloc[-1],  # Use the last MACD signal value as a starting point
        'bollinger_upper': df['bollinger_upper'].iloc[-1],  # Use the last Bollinger upper band value as a starting point
        'bollinger_lower': df['bollinger_lower'].iloc[-1],  # Use the last Bollinger lower band value as a starting point
        'sentiment_score': df['sentiment_score'].iloc[-1],  # Use the last sentiment score as a starting point
        'on_chain_score': df['on_chain_score'].iloc[-1]  # Use the last on-chain score as a starting point
    })
    
    # Make future predictions
    future_predictions = model.predict(future_features)
    
    # Append future predictions to the original DataFrame
    future_df = pd.DataFrame({
        'timestamp': future_features['timestamp'],
        'predicted_close': future_predictions
    })
    
    return df, future_df

def generate_report(symbols, all_data):
    """
    Generate a textual report of the analysis for each symbol.
    """
    report = ""
    for symbol, data in zip(symbols, all_data):
        recommendation = data['recommendation'].iloc[-1]
        final_score = data['final_score'].iloc[-1]
        last_price = data['close'].iloc[-1]
        last_rsi = data['rsi'].iloc[-1]
        last_macd = data['macd'].iloc[-1]
        last_macd_signal = data['macdsignal'].iloc[-1]

        last_price = float(last_price)

        report += f"Symbol: {symbol}\n"
        report += f"Latest Price: {last_price:.2f}\n"
        report += f"Final Score: {final_score:.2f}\n"
        report += f"RSI: {last_rsi:.2f}, MACD: {last_macd:.2f}, MACD Signal: {last_macd_signal:.2f}\n"
        report += f"Recommendation: {recommendation}\n"
        report += "-"*50 + "\n"

    return report

def format_excel_file(filename):
    """
    Apply formatting to the Excel file based on the recommendations.
    """
    wb = load_workbook(filename)
    ws = wb.active

    # Format headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Apply background color based on recommendation
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        recommendation = row[-1].value
        if recommendation == 'Buy':
            for cell in row:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif recommendation == 'Sell':
            for cell in row:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif recommendation == 'Hold':
            for cell in row:
                cell.fill = PatternFill(start_color="FFFC99", end_color="FFFC99", fill_type="solid")

    # Adjust column widths based on content length
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Enable AutoFilter
    ws.auto_filter.ref = ws.dimensions
    wb.save(filename)

def plot_data(all_data, symbols, file_index):
    """
    Plot the data with technical indicators and save all plots to a single image file.
    """
    plt.figure(figsize=(14, 7 * len(symbols)))
    
    for i, (df, symbol) in enumerate(zip(all_data, symbols)):
        if 'rsi' in df.columns and 'macd' in df.columns and 'macdsignal' in df.columns:
            plt.subplot(len(symbols), 1, i + 1)
            plt.plot(df['timestamp'], df['close'], label='Close Price')
            plt.plot(df['timestamp'], df['rsi'], label='RSI')
            plt.plot(df['timestamp'], df['macd'], label='MACD')
            plt.plot(df['timestamp'], df['macdsignal'], label='MACD Signal')
            plt.fill_between(df['timestamp'], df['bollinger_upper'], df['bollinger_lower'], color='gray', alpha=0.2, label='Bollinger Bands')

            # Add buy/sell/hold markers
            buy_signals = df[df['recommendation'] == 'Buy']
            sell_signals = df[df['recommendation'] == 'Sell']
            hold_signals = df[df['recommendation'] == 'Hold']
            
            plt.scatter(buy_signals['timestamp'], buy_signals['close'], marker='^', color='g', label='Buy', s=100)
            plt.scatter(sell_signals['timestamp'], sell_signals['close'], marker='v', color='r', label='Sell', s=100)
            plt.scatter(hold_signals['timestamp'], hold_signals['close'], marker='o', color='b', label='Hold', s=100)

            # Add intermittent data labels for close price
            for idx in range(0, len(df), max(1, len(df) // 10)):  # Adjust the step size for labels
                timestamp = df['timestamp'].iloc[idx]
                close_price = float(df['close'].iloc[idx])
                plt.text(timestamp, close_price, f"{close_price:.2f}", fontsize=8, ha='right')

            plt.title(f'Price and Indicators for {symbol}')
            plt.xlabel('Date')
            plt.ylabel('Value')
            plt.legend()
            plt.grid()
        else:
            print(f"Skipping plot for {symbol} due to missing technical indicators.")
    
    plt.tight_layout()
    plt.savefig(f'all_symbols_analysis_{file_index}.png')
    plt.close()

def on_run():
    """
    Trigger the analysis based on user input from the GUI.
    """
    symbols = [symbol_listbox.get(i) for i in symbol_listbox.curselection()]
    period = period_combobox.get()
    interval = interval_combobox.get()
    
    # Update progress bar
    progress_bar.start()
    
    # Run analysis in a separate thread to keep the GUI responsive
    threading.Thread(target=run_analysis, args=(symbols, period, interval)).start()

def run_analysis(symbols, period, interval):
    """
    Run the analysis for the given symbols, period, and interval.
    """
    all_data = []
    file_index = 1
    for i, symbol in enumerate(symbols):
        price_data = get_price_data(symbol, period, interval)
        if price_data.empty:
            print(f"No data found for {symbol}")
            continue

        sentiment_score = get_sentiment_score(symbol)
        on_chain_metrics = get_on_chain_metrics(symbol)

        price_data = calculate_technical_indicators(price_data)
        price_data = calculate_scores(price_data, sentiment_score, on_chain_metrics)
        all_data.append((price_data, symbol))

        if (i + 1) % 4 == 0 or (i + 1) == len(symbols):
            plot_data([data[0] for data in all_data], [data[1] for data in all_data], file_index)
            report = generate_report([data[1] for data in all_data], [data[0] for data in all_data])
            print(report)

            with pd.ExcelWriter(f'all_symbols_analysis_{file_index}.xlsx') as writer:
                for data, symbol in all_data:
                    if isinstance(data.columns, pd.MultiIndex):
                        data.columns = [' '.join(col).strip() for col in data.columns.values]
                    data.to_excel(writer, sheet_name=symbol, index=False)

            format_excel_file(f'all_symbols_analysis_{file_index}.xlsx')
            print(f"Analysis complete. Results saved to 'all_symbols_analysis_{file_index}.xlsx'")
            messagebox.showinfo("Analysis Complete", f"Results saved to 'all_symbols_analysis_{file_index}.xlsx'")

            all_data = []
            file_index += 1
    
    # Stop progress bar when analysis is done
    progress_bar.stop()

    # Train predictive model and make predictions
    for data, symbol in all_data:
        model = train_predictive_model(data)
        df, future_df = make_predictions(model, data, future_periods=90)  # Make predictions for 90 days
        
        # Plot predictions
        plt.figure(figsize=(14, 7))
        plt.plot(df['timestamp'], df['close'], label='Actual Close Price')
        plt.plot(df['timestamp'], df['predicted_close'], label='Predicted Close Price', linestyle='--')
        plt.plot(future_df['timestamp'], future_df['predicted_close'], label='Future Predicted Close Price', linestyle='--', color='red')
        plt.title(f'Actual vs Predicted Close Price for {symbol}')
        plt.xlabel('Date')
        plt.ylabel('Close Price')
        plt.legend()
        plt.grid()
        plt.savefig(f'predictions_{symbol}_{file_index}.png')
        plt.close()

# Tkinter GUI setup
root = tk.Tk()
root.title("Cryptocurrency Analysis")
root.geometry("600x500")

frame = ttk.Frame(root, padding="10")
frame.grid(row=0, column=0, sticky="nsew")

style = ttk.Style()
style.configure("TLabel", font=("Helvetica", 12))
style.configure("TButton", font=("Helvetica", 12))
style.configure("TCombobox", font=("Helvetica", 12))
style.configure("TListbox", font=("Helvetica", 12))

symbols_label = ttk.Label(frame, text="Symbols (select from list):")
symbols_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
symbol_listbox = tk.Listbox(frame, selectmode=tk.MULTIPLE, height=10, exportselection=0)
symbol_listbox.grid(row=0, column=1, padx=5, pady=5, sticky="w")

# Populate listbox
#crypto_symbols = ['BTC-USD', 'ETH-USD', 'XRP-USD', 'LTC-USD', 'BCH-USD', 'ADA-USD', 'DOT-USD', 'BNB-USD', 'SOL-USD', 'DOGE-USD']
crypto_symbols = ['DOGE-USD']
for symbol in crypto_symbols:
    symbol_listbox.insert(tk.END, symbol)

symbols_label_tooltip = tk.Label(frame, text="Select the cryptocurrencies you want to analyze.")
symbols_label_tooltip.grid(row=0, column=2, padx=5, pady=5, sticky="w")

period_label = ttk.Label(frame, text="Period:")
period_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
period_combobox = ttk.Combobox(frame, values=['1d', '1wk', '1mo', '3mo', '1y'], state="readonly")
period_combobox.grid(row=1, column=1, padx=5, pady=5, sticky="w")
period_combobox.current(4)

interval_label = ttk.Label(frame, text="Interval:")
interval_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
interval_combobox = ttk.Combobox(frame, values=['1m', '5m', '15m', '1h', '1d'], state="readonly")
interval_combobox.grid(row=2, column=1, padx=5, pady=5, sticky="w")
interval_combobox.current(4)

run_button = ttk.Button(frame, text="Run Analysis", command=on_run)
run_button.grid(row=3, column=0, columnspan=2, pady=10)

progress_bar = ttk.Progressbar(frame, orient="horizontal", length=300, mode="determinate")
progress_bar.grid(row=4, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()